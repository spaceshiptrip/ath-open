"""
generate_pairings.py — Generate Blue Crew (Team B / JPL Team) pairings schedule.

Run from repo root:
    .venv/bin/python3 backend/generate_pairings.py

Outputs: docs/BlueCrew_Pairings.xlsx

── ROSTER ────────────────────────────────────────────────────────────────────
  Women (4):  Cora (captain), Alexis, Carmela, Ivy
  Men (9):    Jay, Marv, Arman, Jon, Trevor, Richard, Rhon, Joe, Pierre
  Couples:    Trevor↔Alexis, Marv↔Carmela, Pierre↔Ivy
  Captain pair: Jay+Cora — Jay is Cora's mixed doubles partner (not a couple)

── MIXED DOUBLES DISTRIBUTION ────────────────────────────────────────────────
  Cora (captain):  1 game  (with Jay)
  Alexis:          2 games (with Trevor)
  Carmela:         2 games (with Marv)
  Ivy:             2 games (with Pierre)
  Total: 1+2+2+2 = 7 mixed slots ✓

── HARD CONSTRAINTS (all enforced) ─────────────────────────────────────────
  1. Couples always play together in mixed (fixed-couple iteration)
  2. No player on two courts of the same round
  3. No two games in consecutive rounds for any player (no back-to-back)
  4. No couple/pair assigned to adjacent rounds (R4+R5 is adjacent)
  5. Couple men / Jay don't play men's doubles in the same round as their
     mixed game, or in a round adjacent to their mixed game
  6. No man repeats a men's doubles partner

── MEN'S DOUBLES MATH ────────────────────────────────────────────────────────
  22 men's player-slots total (11 matches × 2)
  Jay:    1 mixed  → 2 men's  (3 total)
  Trevor: 2 mixed  → 1 men's  (3 total)
  Marv:   2 mixed  → 1 men's  (3 total)
  Pierre: 2 mixed  → 1 men's  (3 total)
  Couple/Jay contribution to men's: 2+1+1+1 = 5
  Remaining for non-mixed men (Arman, Jon, Richard, Rhon, Joe): 22-5 = 17
  17 ÷ 5 men = 3 remainder 2 → exactly 2 of the 5 play 4 games

── HOW TO TWEAK ─────────────────────────────────────────────────────────────
  RANDOM_SEED      — integer for reproducible draw, None for random each run
  FORCED_QUAD_MEN  — list of 2 player IDs (from non-mixed pool) who play 4
                     games. e.g. ['b4','b7'] for Arman+Richard. None=random.
  PINNED_PAIRS     — force specific men's doubles pairings, e.g. [('b2','b5')]
"""

import random, sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ════════════════════════════════════════════════════════════════════════════════

RANDOM_SEED     = None   # e.g. 42 for reproducible draw
FORCED_QUAD_MEN = None   # e.g. ['b4','b7'] to pin who plays 4 games
PINNED_PAIRS    = []     # e.g. [('b2','b5')] to force specific men's doubles pairings

# ════════════════════════════════════════════════════════════════════════════════
# ROSTER & SCHEDULE DATA
# ════════════════════════════════════════════════════════════════════════════════

# Hill Street Blues (Team A) — fixed from docs/Team Round Robin Athenaeum.docx
A_SCHED = {
    (1,'S'): ('Pierre',  'Jeff E',   False),
    (1,'N'): ('Dro',     'Steve',    False),
    (2,'S'): ('Mich',    'Wilfred',  False),
    (2,'N'): ('Jeff W',  'Yu Fon',   False),
    (3,'S'): ('Steve',   'Jeff W',   False),
    (3,'N'): ('Dro',     'Johnny',   False),
    (4,'S'): ('Pierre',  'Suzan',    True),
    (4,'N'): ('Jeff E',  'Rachel',   True),
    (5,'S'): ('Jeff E',  'Molly',    True),
    (5,'N'): ('Johnny',  'Wilfred',  False),
    (6,'S'): ('Yu Fon',  'Steve',    False),
    (6,'N'): ('Pierre',  'Mich',     False),
    (7,'S'): ('Johnny',  'Rachel',   True),
    (7,'N'): ('Mich',    'Molly',    True),
    (8,'S'): ('Jeff W',  'Dro',      False),
    (8,'N'): ('Yu Fon',  'Jeff E',   False),
    (9,'S'): ('Wilfred', 'Rachel',   True),
    (9,'N'): ('Pierre',  'Molly',    True),
}

TIMES = {1:'8:30 AM', 2:'8:50 AM', 3:'9:15 AM', 4:'9:40 AM', 5:'10:10 AM',
         6:'10:30 AM', 7:'11:00 AM', 8:'11:30 AM', 9:'—'}

# Blue Crew roster.  b13=Cora added as 4th woman/captain.
PB = {
    'b13': {'name': 'Cora',    'gender': 'F', 'partner': ''},    # captain
    'b1':  {'name': 'Alexis',  'gender': 'F', 'partner': 'b6'},
    'b9':  {'name': 'Carmela', 'gender': 'F', 'partner': 'b3'},
    'b10': {'name': 'Ivy',     'gender': 'F', 'partner': 'b12'},
    'b2':  {'name': 'Jay',     'gender': 'M', 'partner': ''},
    'b3':  {'name': 'Marv',    'gender': 'M', 'partner': 'b9'},
    'b4':  {'name': 'Arman',   'gender': 'M', 'partner': ''},
    'b5':  {'name': 'Jon',     'gender': 'M', 'partner': ''},
    'b6':  {'name': 'Trevor',  'gender': 'M', 'partner': 'b1'},
    'b7':  {'name': 'Richard', 'gender': 'M', 'partner': ''},
    'b8':  {'name': 'Rhon',    'gender': 'M', 'partner': ''},
    'b11': {'name': 'Joe',     'gender': 'M', 'partner': ''},
    'b12': {'name': 'Pierre',  'gender': 'M', 'partner': 'b10'},
}

nm = lambda pid: PB[pid]['name']

# Regular couples (man_id, woman_id) — each plays 2 mixed games
COUPLES = [('b6','b1'), ('b3','b9'), ('b12','b10')]

# Captain pair (man_id, woman_id) — plays exactly 1 mixed game
CAPTAIN_PAIR = ('b2', 'b13')   # Jay + Cora

# All men who appear in mixed (either as couple or captain-pair man)
MIXED_MEN = {m for m,w in COUPLES} | {CAPTAIN_PAIR[0]}

# The 7 mixed doubles slots (round, court)
MIXED_SLOTS = [(4,'S'),(4,'N'),(5,'S'),(7,'S'),(7,'N'),(9,'S'),(9,'N')]

# ════════════════════════════════════════════════════════════════════════════════
# STEP 1 — Initialise RNG
# ════════════════════════════════════════════════════════════════════════════════

if RANDOM_SEED is not None:
    random.seed(RANDOM_SEED)
else:
    random.seed()

# ════════════════════════════════════════════════════════════════════════════════
# STEP 2 — Assign pairs to mixed slots
#           Cora+Jay get exactly 1 slot.
#           Each couple gets exactly 2 slots.
#           Constraints:
#             - No pair in two courts of the same round
#             - No pair in two adjacent rounds (no back-to-back)
# ════════════════════════════════════════════════════════════════════════════════

def assign_mixed_slots():
    """
    Backtracking: place 4 pairs into 7 slots.
    Pair quotas: captain_pair=1, each couple=2.
    No pair in same round or adjacent round twice.
    Returns (slot→pair_idx dict) where 0=captain_pair, 1/2/3=COUPLES[0/1/2].
    """
    all_pairs = [CAPTAIN_PAIR] + COUPLES           # index 0=Cora+Jay, 1-3=couples
    counts    = [1, 2, 2, 2]                       # how many slots each gets

    tickets = []
    for i, c in enumerate(counts):
        tickets.extend([i] * c)

    slots = list(MIXED_SLOTS)
    random.shuffle(slots)

    assignment = {}                    # slot → pair_idx
    pair_rounds = {i: [] for i in range(4)}   # pair_idx → list of rounds used

    def back_to_back(rounds_used, new_r):
        return any(abs(new_r - r) <= 1 for r in rounds_used)

    def bt(s_idx, rem):
        if s_idx == len(slots):
            return True
        slot = slots[s_idx]
        r    = slot[0]
        tried = set()
        candidates = list(range(len(rem)))
        random.shuffle(candidates)
        for i in candidates:
            pi = rem[i]
            if pi in tried:
                continue
            tried.add(pi)
            if r in pair_rounds[pi]:
                continue                      # same round — both courts at once
            if back_to_back(pair_rounds[pi], r):
                continue                      # adjacent round — back-to-back
            pair_rounds[pi].append(r)
            assignment[slot] = pi
            new_rem = rem[:i] + rem[i+1:]
            if bt(s_idx + 1, new_rem):
                return True
            assignment.pop(slot)
            pair_rounds[pi].pop()
        return False

    return assignment if bt(0, tickets) else None

mixed_assignment = None
for _ in range(200):
    mixed_assignment = assign_mixed_slots()
    if mixed_assignment:
        break

if not mixed_assignment:
    print("ERROR: Could not assign mixed slots with no-back-to-back constraint.")
    sys.exit(1)

# Build mixed schedule: slot → (man_name, woman_name, True)
all_pairs_list = [CAPTAIN_PAIR] + COUPLES
mixed_sched = {slot: all_pairs_list[pi] for slot, pi in mixed_assignment.items()}

# For each mixed man, which rounds do they play mixed?
mixed_man_rounds = {}   # man_pid → set of rounds
for slot, (mid, wid) in mixed_sched.items():
    mixed_man_rounds.setdefault(mid, set()).add(slot[0])

# ════════════════════════════════════════════════════════════════════════════════
# STEP 3 — Calculate each man's men's doubles game count
# ════════════════════════════════════════════════════════════════════════════════

mens_ct = {}
for mid, wid in COUPLES:
    mens_ct[mid] = 1   # couple men each play 2 mixed → 1 men's (3 total)
mens_ct[CAPTAIN_PAIR[0]] = 2   # Jay: 1 mixed → 2 men's (3 total)

non_mixed_men = [pid for pid, p in PB.items()
                 if p['gender'] == 'M' and pid not in MIXED_MEN]
# non_mixed_men = Arman, Jon, Richard, Rhon, Joe (5 men)

total_needed = 22 - sum(mens_ct.values())   # 22 - (1+1+1+2) = 17
extras_needed = total_needed - 3 * len(non_mixed_men)   # 17 - 15 = 2

if FORCED_QUAD_MEN is not None:
    extra_men = set(FORCED_QUAD_MEN)
else:
    extra_men = set(random.sample(non_mixed_men, extras_needed))

for pid in non_mixed_men:
    mens_ct[pid] = 4 if pid in extra_men else 3

# ════════════════════════════════════════════════════════════════════════════════
# STEP 4 — Generate valid men's doubles pairs (no repeat partners)
# ════════════════════════════════════════════════════════════════════════════════

def make_pairs(counts, pinned):
    remaining = dict(counts)
    for p1, p2 in pinned:
        remaining[p1] -= 1
        remaining[p2] -= 1

    pool = []
    for pid, c in sorted(remaining.items(), key=lambda x: -x[1]):
        if c > 0:
            pool.extend([pid] * c)

    used   = {(min(a,b), max(a,b)) for a,b in pinned}
    result = list(pinned)

    def bt(rem):
        if not rem:
            return True
        freq = Counter(rem)
        p1   = freq.most_common(1)[0][0]
        others = list({p for p in rem if p != p1})
        random.shuffle(others)
        for p2 in others:
            key = (min(p1,p2), max(p1,p2))
            if key not in used:
                nr = list(rem); nr.remove(p1); nr.remove(p2)
                used.add(key); result.append((p1,p2))
                if bt(nr): return True
                result.pop(); used.discard(key)
        return False

    return result if bt(pool) else None

pairs = None
for _ in range(50):
    pairs = make_pairs(mens_ct, PINNED_PAIRS)
    if pairs:
        break

if not pairs:
    print("ERROR: Could not generate valid men's doubles pairs.")
    sys.exit(1)

# ════════════════════════════════════════════════════════════════════════════════
# STEP 5 — Assign pairs to men's doubles slots
#           Constraints:
#             - No player in two courts of same round
#             - No player in two adjacent rounds (no back-to-back)
#             - Mixed men not in same round OR adjacent round as their mixed game
# ════════════════════════════════════════════════════════════════════════════════

mens_slots = [(r,c) for r in range(1,10)
              for c in ('S','N') if (r,c) not in set(MIXED_SLOTS)]

def blocked_rounds(pid):
    """Rounds where pid cannot play men's doubles."""
    mr = mixed_man_rounds.get(pid, set())
    blocked = set(mr)
    for r in mr:
        blocked.add(r - 1)
        blocked.add(r + 1)
    return blocked

def assign_pairs_to_rounds(pairs, slots):
    shuffled = list(pairs)
    random.shuffle(shuffled)

    assignment  = {}
    player_rounds = {}   # pid → set of rounds played so far

    def can_play(pid, r):
        used = player_rounds.get(pid, set())
        if r in used: return False                         # same round
        if any(abs(r - u) <= 1 for u in used): return False  # back-to-back
        if r in blocked_rounds(pid): return False          # mixed conflict
        return True

    def bt(s_idx, rem):
        if s_idx == len(slots):
            return True
        slot = slots[s_idx]
        r    = slot[0]
        candidates = list(range(len(rem)))
        random.shuffle(candidates)
        for i in candidates:
            p1, p2 = rem[i]
            if can_play(p1, r) and can_play(p2, r):
                player_rounds.setdefault(p1, set()).add(r)
                player_rounds.setdefault(p2, set()).add(r)
                assignment[slot] = (p1, p2)
                new_rem = rem[:i] + rem[i+1:]
                if bt(s_idx + 1, new_rem):
                    return True
                assignment.pop(slot)
                player_rounds[p1].discard(r)
                player_rounds[p2].discard(r)
        return False

    return assignment if bt(0, shuffled) else None

mens_assignment = None
for _ in range(100):
    mens_assignment = assign_pairs_to_rounds(pairs, mens_slots)
    if mens_assignment:
        break

if not mens_assignment:
    print("ERROR: Could not assign pairs to rounds without back-to-back conflicts.")
    print("       Try running again, or relax constraints via PINNED_PAIRS / FORCED_QUAD_MEN.")
    sys.exit(1)

# ════════════════════════════════════════════════════════════════════════════════
# STEP 6 — Build final schedule & partnership index
# ════════════════════════════════════════════════════════════════════════════════

B_SCHED = {}
for slot, (p1, p2) in mens_assignment.items():
    B_SCHED[slot] = (nm(p1), nm(p2), False)
for slot, (mid, wid) in mixed_sched.items():
    B_SCHED[slot] = (nm(mid), nm(wid), True)

b_partnerships = []
for slot in sorted(B_SCHED.keys()):
    p1, p2, is_mix = B_SCHED[slot]
    b_partnerships.append((p1, p2, f"R{slot[0]}-{slot[1]}", is_mix))

pair_map_b = {}
for p1, p2, rnd, mix in b_partnerships:
    pair_map_b[(p1,p2)] = (rnd, mix)
    pair_map_b[(p2,p1)] = (rnd, mix)

b_players_ordered = (
    [pid for pid,p in PB.items() if p['gender']=='F'] +
    [pid for pid,p in PB.items() if p['gender']=='M']
)

game_totals = {}
for pid in b_players_ordered:
    p    = PB[pid]
    name = p['name']
    mix_c = sum(1 for p1n,p2n,_,mix in b_partnerships if mix and (p1n==name or p2n==name))
    mns_c = sum(1 for p1n,p2n,_,mix in b_partnerships if not mix and (p1n==name or p2n==name))
    game_totals[pid] = mix_c + mns_c

# ════════════════════════════════════════════════════════════════════════════════
# STEP 7 — Print summary & validate
# ════════════════════════════════════════════════════════════════════════════════

print(f"\n{'='*68}")
print(f"  BLUE CREW (JPL Team) — Pairings Generated")
print(f"{'='*68}")

print(f"\n  Mixed doubles assignment:")
for slot in sorted(MIXED_SLOTS):
    mid, wid = mixed_sched[slot]
    cap = " ← captain" if mid == CAPTAIN_PAIR[0] else ""
    print(f"    R{slot[0]}{slot[1]}: {nm(mid):8s} + {nm(wid)}{cap}")

print(f"\n  {'Player':10} {'Mix':4} {'Mens':5} {'Total':6} {'Games (round)'}")
print(f"  {'-'*68}")

for pid in b_players_ordered:
    p    = PB[pid]
    name = p['name']
    mix_c = sum(1 for p1n,p2n,_,mix in b_partnerships if mix and (p1n==name or p2n==name))
    mns_c = sum(1 for p1n,p2n,_,mix in b_partnerships if not mix and (p1n==name or p2n==name))
    total = mix_c + mns_c
    partner_strs = []
    for p1n,p2n,rnd,is_mix in b_partnerships:
        opp = p2n if p1n==name else (p1n if p2n==name else None)
        if opp:
            partner_strs.append(f"{opp}@{rnd}{'★' if is_mix else ''}")
    flag = " ← 4 games!" if total==4 else (" ← captain" if pid=='b13' else "")
    print(f"  {name:10} {mix_c:<4} {mns_c:<5} {total:<6} {', '.join(partner_strs)}{flag}")

# Validate
errors = []
for r in range(1,10):
    players_this_round = []
    for c in ('S','N'):
        if (r,c) in B_SCHED:
            p1,p2,_ = B_SCHED[(r,c)]
            players_this_round.extend([p1,p2])
    if len(players_this_round) != len(set(players_this_round)):
        errors.append(f"SAME-ROUND CONFLICT: R{r} — {players_this_round}")

# back-to-back check
from collections import defaultdict
player_round_list = defaultdict(list)
for slot in sorted(B_SCHED.keys()):
    p1,p2,_ = B_SCHED[slot]
    player_round_list[p1].append(slot[0])
    player_round_list[p2].append(slot[0])
for name, rlist in player_round_list.items():
    rlist_s = sorted(set(rlist))
    for i in range(len(rlist_s)-1):
        if rlist_s[i+1] - rlist_s[i] == 1:
            errors.append(f"BACK-TO-BACK: {name} plays R{rlist_s[i]} and R{rlist_s[i+1]}")

mens_partner_counts = Counter()
for p1n,p2n,_,mix in b_partnerships:
    if not mix:
        mens_partner_counts[tuple(sorted([p1n,p2n]))] += 1
for key,cnt in mens_partner_counts.items():
    if cnt > 1:
        errors.append(f"REPEAT PARTNER: {key[0]} + {key[1]} in men's doubles {cnt}×")

if errors:
    print(f"\n  !! VALIDATION ERRORS:")
    for e in errors: print(f"     {e}")
    sys.exit(1)
else:
    print(f"\n  ✓ No same-round conflicts · No back-to-back games · No repeated partners")

print(f"\n  (2 men play 4 games — unavoidable with 17 slots / 5 non-mixed men)")

# ════════════════════════════════════════════════════════════════════════════════
# XLSX STYLES
# ════════════════════════════════════════════════════════════════════════════════

G_DARK   = PatternFill("solid", fgColor="1d4f35")
G_MID    = PatternFill("solid", fgColor="2d7d4f")
G_LIGHT  = PatternFill("solid", fgColor="C8E6C9")
YELLOW   = PatternFill("solid", fgColor="FFF176")
GREY     = PatternFill("solid", fgColor="E0E0E0")
RED_L    = PatternFill("solid", fgColor="FFDCE0")
BLUE_L   = PatternFill("solid", fgColor="DCE8FF")
WHITE    = PatternFill("solid", fgColor="FFFFFF")
ORANGE_L = PatternFill("solid", fgColor="FFE0B2")
PURPLE_L = PatternFill("solid", fgColor="EDE7F6")

thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cs(c, fill=None, bold=False, color="000000", align="center", wrap=False, size=10):
    if fill: c.fill = fill
    c.font = Font(bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = border

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Full Schedule
# ════════════════════════════════════════════════════════════════════════════════

wb = Workbook()
ws1 = wb.active
ws1.title = "Full Schedule"

ws1.merge_cells("A1:I1")
t = ws1["A1"]
t.value = "ATH Open — Full Tournament Schedule  (Hill Street Blues  vs  Blue Crew / JPL Team)"
cs(t, fill=G_DARK, bold=True, color="FFFFFF", size=12)
ws1.row_dimensions[1].height = 26

hdrs = ["Round","Time","Court","Type","Hill Street Blues","","Blue Crew (JPL)","","Notes"]
sub  = ["","","","","Player 1","Player 2","Player 1","Player 2",""]
for ci,h in enumerate(hdrs,1):
    c = ws1.cell(row=2, column=ci, value=h)
    cs(c, fill=G_MID, bold=True, color="FFFFFF", size=9)
for ci,h in enumerate(sub,1):
    c = ws1.cell(row=3, column=ci, value=h)
    cs(c, fill=G_DARK, bold=True, color="FFFFFF", size=8)
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 16

for r in range(1,10):
    for court in ('S','N'):
        slot   = (r,court)
        a1,a2,is_mix = A_SCHED[slot]
        b1_,b2_,_    = B_SCHED[slot]
        row = 4 + (r-1)*2 + (0 if court=='S' else 1)
        type_str   = "Mix Doubles ★" if is_mix else "Men's Doubles"
        court_str  = f"{court} ({'South' if court=='S' else 'North'})"
        # Flag if this is the captain's game
        is_cap     = slot in mixed_sched and mixed_sched[slot][0] == CAPTAIN_PAIR[0]
        note       = "Captain's game ★" if is_cap else ""
        vals = [r, TIMES[r], court_str, type_str, a1, a2, b1_, b2_, note]
        for ci,val in enumerate(vals,1):
            c = ws1.cell(row=row, column=ci, value=val)
            c.fill = YELLOW if is_mix else WHITE
            c.font = Font(bold=is_mix, size=10)
            c.alignment = Alignment(
                horizontal="center" if ci not in (5,6,7,8,9) else "left",
                vertical="center")
            c.border = border
        ws1.row_dimensions[row].height = 18

for i,w in enumerate([7,10,14,16,14,14,14,14,16],1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Blue Crew Pairings Matrix
# ════════════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Blue Crew Matrix")

b_names_ordered = (
    [nm(pid) for pid,p in PB.items() if p['gender']=='F'] +
    [nm(pid) for pid,p in PB.items() if p['gender']=='M']
)
b_gender = {nm(pid): p['gender'] for pid,p in PB.items()}

n = len(b_names_ordered)
START_COL  = 3
HEADER_ROW = 4
GAMES_COL  = START_COL + n
MIX_COL    = START_COL + n + 1

ws2.merge_cells(f"A1:{get_column_letter(MIX_COL+1)}1")
t2 = ws2["A1"]
t2.value = "Blue Crew (JPL Team) — Pairings Matrix   (X = Men's Doubles  |  M★ = Mixed Doubles  |  C★ = Captain's game)"
cs(t2, fill=G_DARK, bold=True, color="FFFFFF", size=11)
ws2.row_dimensions[1].height = 22

ws2.merge_cells(f"A2:{get_column_letter(MIX_COL+1)}2")
leg2 = ws2["A2"]
leg2.value = "Captain: Cora (1 game w/ Jay)  ·  Couples: Trevor↔Alexis, Marv↔Carmela, Pierre↔Ivy  |  No back-to-back games"
cs(leg2, fill=G_MID, bold=False, color="FFFFFF", size=9)
ws2.row_dimensions[2].height = 16
ws2.row_dimensions[3].height = 6

for col,label in [(1,"Rank"),(2,"Player")]:
    c = ws2.cell(row=HEADER_ROW, column=col, value=label)
    cs(c, fill=G_DARK, bold=True, color="FFFFFF", size=9)

for ci,pname in enumerate(b_names_ordered):
    col    = START_COL + ci
    gender = b_gender[pname]
    is_cap = pname == 'Cora'
    label  = f"{pname}\n({'C★' if is_cap else ('F' if gender=='F' else 'M')})"
    c = ws2.cell(row=HEADER_ROW, column=col, value=label)
    c.fill = PURPLE_L if is_cap else (RED_L if gender=='F' else BLUE_L)
    c.font = Font(bold=True, size=8)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws2.column_dimensions[get_column_letter(col)].width = 7

for col,label in [(GAMES_COL,"Games"),(MIX_COL,"Mix★")]:
    c = ws2.cell(row=HEADER_ROW, column=col, value=label)
    cs(c, fill=G_DARK, bold=True, color="FFFFFF", size=9)
ws2.row_dimensions[HEADER_ROW].height = 36

for ri,pname in enumerate(b_names_ordered):
    row    = HEADER_ROW + 1 + ri
    gender = b_gender[pname]
    is_cap = pname == 'Cora'
    row_fill = PURPLE_L if is_cap else (RED_L if gender=='F' else BLUE_L)
    cs(ws2.cell(row=row, column=1, value=ri+1), fill=row_fill, bold=True, size=9)
    cs(ws2.cell(row=row, column=2, value=f"{pname}{' (Captain)' if is_cap else ''}"),
       fill=row_fill, bold=True, align="left", size=9)
    for ci,oname in enumerate(b_names_ordered):
        col = START_COL + ci
        c   = ws2.cell(row=row, column=col)
        if pname == oname:
            c.fill = GREY; c.border = border
        elif (pname,oname) in pair_map_b:
            rnd,mix = pair_map_b[(pname,oname)]
            is_cap_game = mix and (pname=='Cora' or oname=='Cora')
            c.value = f"C★\n{rnd}" if is_cap_game else (f"M★\n{rnd}" if mix else f"X\n{rnd}")
            c.fill  = PURPLE_L if is_cap_game else (YELLOW if mix else G_LIGHT)
            c.font  = Font(bold=True, size=8,
                           color=("5E2CA5" if is_cap_game else ("7B4F00" if mix else "1B5E20")))
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        else:
            c.fill = WHITE; c.border = border
    total  = game_totals[next(k for k,v in PB.items() if v['name']==pname)]
    mix_c  = sum(1 for p1n,p2n,_,mix in b_partnerships if mix and (p1n==pname or p2n==pname))
    cs(ws2.cell(row=row, column=GAMES_COL, value=total),
       fill=ORANGE_L if total==4 else row_fill, bold=True, size=10)
    cs(ws2.cell(row=row, column=MIX_COL,   value=mix_c if mix_c else "—"),
       fill=YELLOW if mix_c else row_fill, bold=bool(mix_c), size=10)
    ws2.row_dimensions[row].height = 24

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 16
ws2.freeze_panes = f"C{HEADER_ROW+1}"

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Player Summary
# ════════════════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("Player Summary")
sum_hdrs = ["#","Player","Gender","Role","Total Games","Men's Dbl","Mixed★","Games in order (round)"]
for ci,h in enumerate(sum_hdrs,1):
    c = ws3.cell(row=1, column=ci, value=h)
    cs(c, fill=G_DARK, bold=True, color="FFFFFF", size=10)
ws3.row_dimensions[1].height = 22

for ri,pid in enumerate(b_players_ordered,2):
    p      = PB[pid]
    name   = p['name']
    gender = p['gender']
    is_cap = pid == 'b13'
    role   = "Captain" if is_cap else ("Couple" if p['partner'] else "—")
    partner_name = nm(p['partner']) if p['partner'] else ("Jay" if is_cap else "—")
    total  = game_totals[pid]
    mix_c  = sum(1 for p1n,p2n,_,mix in b_partnerships if mix and (p1n==name or p2n==name))
    mns_c  = total - mix_c
    ordered = []
    for slot in sorted(B_SCHED.keys()):
        p1n,p2n,mix = B_SCHED[slot]
        if p1n==name or p2n==name:
            opp = p2n if p1n==name else p1n
            ordered.append(f"R{slot[0]}{'★' if mix else ''} w/{opp}")
    row_fill = PURPLE_L if is_cap else (RED_L if gender=='F' else (ORANGE_L if total==4 else BLUE_L))
    vals = [ri-1, name, gender, role, total, mns_c, mix_c if mix_c else 0, ",  ".join(ordered)]
    for ci,val in enumerate(vals,1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.fill = row_fill
        c.font = Font(bold=(ci in (2,5)), size=10)
        c.alignment = Alignment(horizontal="center" if ci!=8 else "left",
                                vertical="center", wrap_text=(ci==8))
        c.border = border
    ws3.row_dimensions[ri].height = 18

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 8
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 12
ws3.column_dimensions["F"].width = 10
ws3.column_dimensions["G"].width = 10
ws3.column_dimensions["H"].width = 80
ws3.freeze_panes = "A2"

wb.save("docs/BlueCrew_Pairings.xlsx")
print(f"\n  Saved → docs/BlueCrew_Pairings.xlsx")
print(f"{'='*68}\n")
