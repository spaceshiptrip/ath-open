"""
Generates backend/ATH_Open_Sheets_Template.xlsx — run from the repo root with:
    .venv/bin/python3 backend/generate_template.py

The output matches the current hardcoded mock data exactly (same player IDs,
same match pairings) so you can paste this directly into Google Sheets and
connect it with code.gs without any changes.

Requires openpyxl (already in .venv):
    .venv/bin/pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── colours ───────────────────────────────────────────────────────────────────
GREEN   = PatternFill("solid", fgColor="1d4f35")
YELLOW  = PatternFill("solid", fgColor="f5c518")
RED     = PatternFill("solid", fgColor="FFDCE0")
BLUE    = PatternFill("solid", fgColor="DCE8FF")
PURPLE  = PatternFill("solid", fgColor="EDE7F6")
GREY    = PatternFill("solid", fgColor="F2F2F2")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD       = Font(bold=True)

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(ws, headers, fill=GREEN, font=WHITE_FONT):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border


def cell(ws, row, col, val, fill=None, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    c.font = Font(bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border
    return c


# ════════════════════════════════════════════════════════════════════════════════
# 1. Players tab — matches src/data/mockData.js exactly
# ════════════════════════════════════════════════════════════════════════════════
ws_p = wb.active
ws_p.title = "Players"
ws_p.row_dimensions[1].height = 28

player_headers = [
    "id", "firstName", "lastName", "team", "gender",
    "isCaptain", "headshotUrl", "phone", "email", "timestamp", "partnerId",
]
hdr(ws_p, player_headers)

col_widths = [14, 12, 10, 6, 8, 10, 36, 16, 28, 28, 12]
for i, w in enumerate(col_widths, 1):
    ws_p.column_dimensions[get_column_letter(i)].width = w

# Hill Street Blues (Team A) — 12 players: 3F + 9M. No fixed couples.
team_a = [
    # id       first     last  team gen  cap    headshot  phone  email  ts    partner
    ["a1",  "Suzan",   "",  "A", "F", "TRUE",  "", "", "", "", ""],
    ["a2",  "Rachel",  "",  "A", "F", "FALSE", "", "", "", "", ""],
    ["a3",  "Molly",   "",  "A", "F", "FALSE", "", "", "", "", ""],
    ["a4",  "Pierre",  "",  "A", "M", "FALSE", "", "", "", "", ""],
    ["a5",  "Jeff",    "E", "A", "M", "FALSE", "", "", "", "", ""],
    ["a6",  "Dro",     "",  "A", "M", "FALSE", "", "", "", "", ""],
    ["a7",  "Steve",   "",  "A", "M", "FALSE", "", "", "", "", ""],
    ["a8",  "Mich",    "",  "A", "M", "FALSE", "", "", "", "", ""],
    ["a9",  "Wilfred", "",  "A", "M", "FALSE", "", "", "", "", ""],
    ["a10", "Jeff",    "W", "A", "M", "FALSE", "", "", "", "", ""],
    ["a11", "Yu Fon",  "",  "A", "M", "FALSE", "", "", "", "", ""],
    ["a12", "Johnny",  "",  "A", "M", "FALSE", "", "", "", "", ""],
]

# Blue Crew (Team B / JPL Team) — 13 players: 4F + 9M
# Captain: Cora (b13) — plays 1 mixed game with Jay (R7S). No fixed partner.
# Fixed couples (play mixed together): Alexis↔Trevor, Carmela↔Marv, Ivy↔Pierre
team_b = [
    ["b13", "Cora",    "", "B", "F", "TRUE",  "", "", "", "", ""],
    ["b1",  "Alexis",  "", "B", "F", "FALSE", "", "", "", "", "b6"],
    ["b9",  "Carmela", "", "B", "F", "FALSE", "", "", "", "", "b3"],
    ["b10", "Ivy",     "", "B", "F", "FALSE", "", "", "", "", "b12"],
    ["b2",  "Jay",     "", "B", "M", "FALSE", "", "", "", "", ""],
    ["b3",  "Marv",    "", "B", "M", "FALSE", "", "", "", "", "b9"],
    ["b4",  "Arman",   "", "B", "M", "FALSE", "", "", "", "", ""],
    ["b5",  "Jon",     "", "B", "M", "FALSE", "", "", "", "", ""],
    ["b6",  "Trevor",  "", "B", "M", "FALSE", "", "", "", "", "b1"],
    ["b7",  "Richard", "", "B", "M", "FALSE", "", "", "", "", ""],
    ["b8",  "Rhon",    "", "B", "M", "FALSE", "", "", "", "", ""],
    ["b11", "Joe",     "", "B", "M", "FALSE", "", "", "", "", ""],
    ["b12", "Pierre",  "Y", "B", "M", "FALSE", "", "", "", "", "b10"],
]

for row_idx, row in enumerate(team_a + team_b, 3):
    is_a = row[3] == "A"
    is_cap = row[5] == "TRUE"
    fill = PURPLE if is_cap and not is_a else (RED if is_a else BLUE)
    for col, val in enumerate(row, 1):
        c = ws_p.cell(row=row_idx, column=col, value=val)
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border
    ws_p.row_dimensions[row_idx].height = 18

ws_p.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════════
# 2. Schedule tab — matches mockData.js exactly, with actual pairings pre-filled
#    Columns: id, round, court, isMix, teamAP1, teamAP2, teamBP1, teamBP2, winner, scoreA, scoreB
# ════════════════════════════════════════════════════════════════════════════════
ws_s = wb.create_sheet("Schedule")
ws_s.row_dimensions[1].height = 28

sched_headers = [
    "id", "round", "court", "isMix",
    "teamAP1", "teamAP2", "teamBP1", "teamBP2",
    "winner", "scoreA", "scoreB",
]
hdr(ws_s, sched_headers)

sched_widths = [6, 7, 7, 8, 10, 10, 10, 10, 8, 8, 8]
for i, w in enumerate(sched_widths, 1):
    ws_s.column_dimensions[get_column_letter(i)].width = w

# Matches exactly matching src/data/mockData.js
# isMix=TRUE rows highlighted in yellow
matches = [
    # id     round court isMix  AP1   AP2   BP1   BP2   win  sA sB
    # R1 8:30 AM — Men's Doubles
    ["m1",  1,"S","FALSE","a4", "a5", "b6", "b8", "", "", ""],
    ["m2",  1,"N","FALSE","a6", "a7", "b7", "b11","", "", ""],
    # R2 8:50 AM — Men's Doubles
    ["m3",  2,"S","FALSE","a8", "a9", "b4", "b12","", "", ""],
    ["m4",  2,"N","FALSE","a10","a11","b3", "b5", "", "", ""],
    # R3 9:15 AM — Men's Doubles
    ["m5",  3,"S","FALSE","a7", "a10","b8", "b11","", "", ""],
    ["m6",  3,"N","FALSE","a6", "a12","b2", "b7", "", "", ""],
    # R4 9:40 AM — Mixed Doubles ★ both courts
    ["m7",  4,"S","TRUE", "a4", "a1", "b12","b10","", "", ""],
    ["m8",  4,"N","TRUE", "a5", "a2", "b3", "b9", "", "", ""],
    # R5 10:10 AM — Mixed South ★, Men's North
    ["m9",  5,"S","TRUE", "a5", "a3", "b6", "b1", "", "", ""],
    ["m10", 5,"N","FALSE","a12","a9", "b2", "b8", "", "", ""],
    # R6 10:30 AM — Men's Doubles
    ["m11", 6,"S","FALSE","a11","a7", "b4", "b7", "", "", ""],
    ["m12", 6,"N","FALSE","a4", "a8", "b5", "b11","", "", ""],
    # R7 11:00 AM — Mixed Doubles ★ both courts
    ["m13", 7,"S","TRUE", "a12","a2", "b2", "b13","", "", ""],
    ["m14", 7,"N","TRUE", "a8", "a3", "b12","b10","", "", ""],
    # R8 11:30 AM — Men's Doubles
    ["m15", 8,"S","FALSE","a10","a6", "b4", "b11","", "", ""],
    ["m16", 8,"N","FALSE","a11","a5", "b5", "b8", "", "", ""],
    # R9 — Mixed Doubles ★ both courts
    ["m17", 9,"S","TRUE", "a9", "a2", "b6", "b1", "", "", ""],
    ["m18", 9,"N","TRUE", "a4", "a3", "b3", "b9", "", "", ""],
]

for row_idx, row in enumerate(matches, 2):
    is_mix = row[3] == "TRUE"
    fill = YELLOW if is_mix else GREY
    for col, val in enumerate(row, 1):
        c = ws_s.cell(row=row_idx, column=col, value=val)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws_s.row_dimensions[row_idx].height = 18

ws_s.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════════════════════
# 3. How To Use tab
# ════════════════════════════════════════════════════════════════════════════════
ws_h = wb.create_sheet("How To Use")
ws_h.column_dimensions["A"].width = 26
ws_h.column_dimensions["B"].width = 72

section_fill = PatternFill("solid", fgColor="2d7d4f")

instructions = [
    ("SETUP", ""),
    ("1. Run initSheets()",  "In the Apps Script editor, select initSheets from the function dropdown and Run. This creates the Players and Schedule tabs and pre-populates all 18 matches. Do this ONCE before the tournament."),
    ("2. Deploy as Web App", "Apps Script → Deploy → New deployment → Web app. Execute as: Me. Who has access: Anyone. Copy the /exec URL."),
    ("3. Set VITE_SHEETS_API_URL", "Paste the /exec URL into GitHub → Settings → Secrets → VITE_SHEETS_API_URL. Re-run the GitHub Actions workflow to rebuild."),
    ("", ""),
    ("PLAYERS TAB", ""),
    ("id",                "Pre-assigned (a1–a12 for HSB, b1–b13 for Blue Crew). Registration form auto-generates p<timestamp> for new players."),
    ("firstName / lastName", "Player's name. lastName can be blank (most players). Set lastName for disambiguation: Jeff E / Jeff W."),
    ("team",              "A (Hill Street Blues) or B (Blue Crew / JPL Team)"),
    ("gender",            "M or F"),
    ("isCaptain",         "TRUE for one player per team; FALSE for all others. Blue Crew captain = Cora (b13)."),
    ("headshotUrl",       "Optional: direct URL to a portrait photo (JPEG or PNG)"),
    ("partnerId",         "The id of this player's fixed mixed-doubles partner. Blank if no fixed partner. Examples: b6 for Alexis (her couple is Trevor), b13 for Cora (blank — Jay is not a fixed couple)."),
    ("phone / email",     "Optional. Not displayed on the site."),
    ("timestamp",         "Auto-filled when a player registers via the web form. Leave blank for pre-loaded players."),
    ("", ""),
    ("SCHEDULE TAB", ""),
    ("id",                "Match IDs m1–m18 are pre-filled. Do not change."),
    ("round / court",     "Pre-filled. Round 1–9, court S (South) or N (North)."),
    ("isMix",             "TRUE = Mixed Doubles ★. Pre-filled. Yellow rows."),
    ("teamAP1/AP2",       "Hill Street Blues player IDs for this match. Pre-filled with the locked HSB pairings from their docx."),
    ("teamBP1/BP2",       "Blue Crew player IDs for this match. Pre-filled with the generated pairings from generate_pairings.py."),
    ("winner",            "Enter A or B after the match is played. Or use the app Scores page — the captain presses A Wins / B Wins on their phone."),
    ("scoreA / scoreB",   "Optional numeric scores. Enter after the match, e.g. 11 and 9. The app shows these alongside the winner badge. The Scores page has numeric input fields in edit mode."),
    ("", ""),
    ("LIVE UPDATES", ""),
    ("How it works",      "Once VITE_SHEETS_API_URL is set, the app reads from this sheet on every page load. No rebuild or redeploy needed — edit this sheet and the website updates immediately."),
    ("Roster changes",    "Add, edit, or remove players in the Players tab. Changes appear on the Teams page within seconds."),
    ("Score entry",       "Enter winner (A/B) and optional scores in the Schedule tab. Changes appear on the Scores and Schedule pages immediately. Captains can also enter scores directly in the app."),
    ("", ""),
    ("PLAYER ID REFERENCE", ""),
    ("Hill Street Blues", "a1=Suzan  a2=Rachel  a3=Molly  a4=Pierre  a5=Jeff E  a6=Dro  a7=Steve  a8=Mich  a9=Wilfred  a10=Jeff W  a11=Yu Fon  a12=Johnny"),
    ("Blue Crew",         "b13=Cora(C)  b1=Alexis  b9=Carmela  b10=Ivy  b2=Jay  b3=Marv  b4=Arman  b5=Jon  b6=Trevor  b7=Richard  b8=Rhon  b11=Joe  b12=Pierre"),
]

for row_idx, (key, val) in enumerate(instructions, 1):
    ck = ws_h.cell(row=row_idx, column=1, value=key)
    cv = ws_h.cell(row=row_idx, column=2, value=val)
    if val == "" and key:
        ck.fill = section_fill
        ck.font = Font(color="FFFFFF", bold=True)
        ws_h.merge_cells(f"A{row_idx}:B{row_idx}")
    else:
        ck.font = Font(bold=True if key else False)
    cv.alignment = Alignment(wrap_text=True, vertical="center")
    ws_h.row_dimensions[row_idx].height = 36 if val and len(val) > 60 else 20


# ════════════════════════════════════════════════════════════════════════════════
# Helper structures for matrix / summary sheets
# ════════════════════════════════════════════════════════════════════════════════

ROUND_TIMES = {
    1: "8:30 AM", 2: "8:50 AM", 3: "9:15 AM", 4: "9:40 AM",
    5: "10:10 AM", 6: "10:30 AM", 7: "11:00 AM", 8: "11:30 AM", 9: "—",
}

# Build player lookup  {id: {id, name, first, last, team, gender, isCaptain, partnerId}}
player_map = {}
for _p in team_a + team_b:
    _pid = _p[0]
    _name = f"{_p[1]} {_p[2]}".strip() if _p[2] else _p[1]
    player_map[_pid] = {
        "id": _pid, "name": _name, "first": _p[1], "last": _p[2],
        "team": _p[3], "gender": _p[4],
        "isCaptain": _p[5] == "TRUE", "partnerId": _p[10],
    }

# Pairings maps: frozenset({id1, id2}) -> list of (round, court, isMix)
hsb_pairs = {}
bc_pairs  = {}
for _m in matches:
    _r, _c, _mix = _m[1], _m[2], _m[3] == "TRUE"
    _ka = frozenset({_m[4], _m[5]})
    hsb_pairs.setdefault(_ka, []).append((_r, _c, _mix))
    _kb = frozenset({_m[6], _m[7]})
    bc_pairs.setdefault(_kb, []).append((_r, _c, _mix))

CAPTAIN_PAIR  = frozenset({"b2", "b13"})   # Jay + Cora, R7-S
FOUR_GAME_IDS = {"b8", "b11"}              # Rhon and Joe play 4 games

# Display order
HSB_ORDER = [player_map[i] for i in
             ["a4","a5","a6","a7","a8","a9","a10","a11","a12","a1","a2","a3"]]
BC_ORDER  = [player_map[i] for i in
             ["b13","b1","b9","b10","b2","b3","b4","b5","b6","b7","b8","b11","b12"]]

# Matrix / summary palette
DK_GRN = PatternFill("solid", fgColor="1d4f35")
MD_GRN = PatternFill("solid", fgColor="2d7d4f")
R_BLUE = PatternFill("solid", fgColor="DCE8FF")
R_PINK = PatternFill("solid", fgColor="FFDCE0")
R_PURP = PatternFill("solid", fgColor="EDE7F6")
R_ORNG = PatternFill("solid", fgColor="FFE0B2")
C_MEN  = PatternFill("solid", fgColor="C8E6C9")
C_MIX  = PatternFill("solid", fgColor="FFF176")
C_CAP  = PatternFill("solid", fgColor="EDE7F6")
C_DIAG = PatternFill("solid", fgColor="E0E0E0")
C_EMPT = PatternFill("solid", fgColor="FFFFFF")
F_MEN  = Font(color="1B5E20", bold=True, size=8)
F_MIX  = Font(color="7B4F00", bold=True, size=8)
F_CAP  = Font(color="5E2CA5", bold=True, size=8)
F_WHT  = Font(color="FFFFFF", bold=True)
F_WHT9 = Font(color="FFFFFF", bold=True, size=9)


def _prow_fill(p, four_ids=None):
    if p["isCaptain"] and p["team"] == "B":
        return R_PURP
    if four_ids and p["id"] in four_ids:
        return R_ORNG
    return R_PINK if p["gender"] == "F" else R_BLUE


def _player_games(pid, pairs):
    """Sorted list of (round, court, isMix, partner_id) for one player."""
    out = []
    for key, games in pairs.items():
        if pid in key:
            partner = next(iter(key - {pid}))
            for (r, c, m) in games:
                out.append((r, c, m, partner))
    return sorted(out)


def _mc(ws, r, c, val, fill=None, fnt=None, align="center", wrap=False):
    """Write a matrix cell."""
    cell_ = ws.cell(row=r, column=c, value=val)
    if fill: cell_.fill = fill
    if fnt:  cell_.font = fnt
    cell_.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell_.border = border
    return cell_


# ════════════════════════════════════════════════════════════════════════════════
# 4. Full Schedule — both teams, all 18 matches with player names
# ════════════════════════════════════════════════════════════════════════════════
ws_fs = wb.create_sheet("Full Schedule")

# Title row
ws_fs.merge_cells("A1:I1")
_mc(ws_fs, 1, 1,
    "ATH Open — Full Tournament Schedule  (Hill Street Blues  vs  Blue Crew / JPL Team)",
    fill=DK_GRN, fnt=Font(color="FFFFFF", bold=True, size=12))
ws_fs.row_dimensions[1].height = 28

# Header row 2
for _ci, _h in enumerate(["Round","Time","Court","Type","Hill Street Blues","","Blue Crew (JPL)","","Notes"], 1):
    _mc(ws_fs, 2, _ci, _h, fill=MD_GRN, fnt=F_WHT9)
ws_fs.row_dimensions[2].height = 20

# Header row 3
for _ci, _h in enumerate(["","","","","Player 1","Player 2","Player 1","Player 2",""], 1):
    _mc(ws_fs, 3, _ci, _h if _h else None, fill=DK_GRN, fnt=F_WHT9)
ws_fs.row_dimensions[3].height = 18

for _mi, _m in enumerate(matches):
    _row = 4 + _mi
    _rnd, _court, _is_mix = _m[1], _m[2], _m[3] == "TRUE"
    _ap1, _ap2, _bp1, _bp2 = _m[4], _m[5], _m[6], _m[7]
    _rfill = C_MIX if _is_mix else C_EMPT
    _rfnt  = Font(bold=_is_mix, size=10)
    _note  = "Captain's game ★" if frozenset({_bp1, _bp2}) == CAPTAIN_PAIR else None
    _vals  = [
        _rnd,
        ROUND_TIMES[_rnd],
        "S (South)" if _court == "S" else "N (North)",
        "Mix Doubles ★" if _is_mix else "Men's Doubles",
        player_map[_ap1]["name"], player_map[_ap2]["name"],
        player_map[_bp1]["name"], player_map[_bp2]["name"],
        _note,
    ]
    for _ci, _v in enumerate(_vals, 1):
        _mc(ws_fs, _row, _ci, _v, fill=_rfill, fnt=_rfnt)
    ws_fs.row_dimensions[_row].height = 18

for _ci, _w in enumerate([7, 10, 14, 16, 14, 14, 14, 14, 18], 1):
    ws_fs.column_dimensions[get_column_letter(_ci)].width = _w
ws_fs.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════════════════════
# Helper: write a pairings matrix sheet
# ════════════════════════════════════════════════════════════════════════════════
def write_matrix(title, subtitle, players, pairs, sheet_name, cap_pair=None, four_ids=None):
    ws = wb.create_sheet(sheet_name)
    n = len(players)
    last_col = 2 + n + 2

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    _mc(ws, 1, 1, title, fill=DK_GRN, fnt=Font(color="FFFFFF", bold=True, size=11))
    ws.row_dimensions[1].height = 28

    # Row 2: subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    _mc(ws, 2, 1, subtitle, fill=MD_GRN, fnt=Font(color="FFFFFF", size=9))
    ws.row_dimensions[2].height = 20

    # Row 3: spacer
    ws.row_dimensions[3].height = 6

    # Row 4: column headers
    _mc(ws, 4, 1, "Rank", fill=DK_GRN, fnt=F_WHT9)
    _mc(ws, 4, 2, "Player", fill=DK_GRN, fnt=F_WHT9)

    for _i, _p in enumerate(players):
        _col = 3 + _i
        _g, _cap, _team = _p["gender"], _p["isCaptain"], _p["team"]
        _nm = _p["name"]
        if _cap and _team == "B":
            _htxt, _hfill = f"{_nm}\n(C★)", R_PURP
        elif _g == "F" and _team == "A":
            _htxt = f"{_nm}\n(F★) ©" if _cap else f"{_nm}\n(F★)"
            _hfill = R_PINK
        elif _g == "F":
            _htxt, _hfill = f"{_nm}\n(F)", R_PINK
        else:
            _htxt, _hfill = f"{_nm}\n(M)", R_BLUE
        _mc(ws, 4, _col, _htxt, fill=_hfill,
            fnt=Font(bold=True, size=8), wrap=True)

    _mc(ws, 4, 3 + n,     "Games", fill=DK_GRN, fnt=F_WHT9)
    _mc(ws, 4, 3 + n + 1, "Mix★",  fill=DK_GRN, fnt=F_WHT9)
    ws.row_dimensions[4].height = 36

    # Rows 5+: player rows
    for _ri, _player in enumerate(players):
        _row = 5 + _ri
        _pid = _player["id"]
        _g, _cap, _team = _player["gender"], _player["isCaptain"], _player["team"]
        _rfill = _prow_fill(_player)  # base row fill (no orange for data rows, only Games col)

        # Rank
        _mc(ws, _row, 1, _ri + 1, fill=_rfill, fnt=Font(size=9))

        # Player label
        if _g == "M":
            _plabel = f"{_player['name']} (M)"
        elif _cap and _team == "A":
            _plabel = f"{_player['name']} (F) ©"
        elif _cap and _team == "B":
            _plabel = f"{_player['name']} (C★) Captain"
        else:
            _plabel = f"{_player['name']} (F)"
        _mc(ws, _row, 2, _plabel, fill=_rfill, fnt=Font(bold=True, size=9), align="left")

        # Compute game counts for this player
        _pg = _player_games(_pid, pairs)
        _total_g = len(_pg)
        _mix_g   = sum(1 for _x in _pg if _x[2])

        # Matrix cells
        for _ci, _opp in enumerate(players):
            _col = 3 + _ci
            _oid = _opp["id"]
            if _pid == _oid:
                _mc(ws, _row, _col, None, fill=C_DIAG)
            else:
                _key = frozenset({_pid, _oid})
                _games = pairs.get(_key, [])
                if not _games:
                    _mc(ws, _row, _col, None, fill=C_EMPT)
                else:
                    _is_cap_game = cap_pair is not None and _key == cap_pair
                    _all_mix = all(_x[2] for _x in _games)
                    _rstr = ",".join(f"R{_x[0]}-{_x[1]}" for _x in _games)
                    if _is_cap_game:
                        _val = f"C★\n{_rstr}"
                        _mc(ws, _row, _col, _val, fill=C_CAP, fnt=F_CAP, wrap=True)
                    elif _all_mix:
                        _val = f"M★\n{_rstr}"
                        _mc(ws, _row, _col, _val, fill=C_MIX, fnt=F_MIX, wrap=True)
                    else:
                        _val = f"X\n{_rstr}"
                        _mc(ws, _row, _col, _val, fill=C_MEN, fnt=F_MEN, wrap=True)

        # Games column (orange if 4-game player)
        _gcol = 3 + n
        _gfill = R_ORNG if (four_ids and _pid in four_ids) else _rfill
        _mc(ws, _row, _gcol, _total_g, fill=_gfill, fnt=Font(bold=True, size=10))

        # Mix★ column
        _mcol = 3 + n + 1
        _mfill = C_MIX if _mix_g > 0 else _rfill
        _mc(ws, _row, _mcol, _mix_g if _mix_g > 0 else "—",
            fill=_mfill, fnt=Font(bold=True, size=10))

        ws.row_dimensions[_row].height = 30

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    for _i in range(n):
        ws.column_dimensions[get_column_letter(3 + _i)].width = 7
    ws.column_dimensions[get_column_letter(3 + n)].width = 8
    ws.column_dimensions[get_column_letter(4 + n)].width = 8
    ws.freeze_panes = "C5"
    return ws


# ════════════════════════════════════════════════════════════════════════════════
# Helper: write a player summary sheet
# ════════════════════════════════════════════════════════════════════════════════
def write_summary(sheet_name, players, pairs, four_ids=None):
    ws = wb.create_sheet(sheet_name)
    _hdrs = ["#", "Player", "Gender", "Role", "Total Games", "Men's Dbl", "Mixed★",
             "Games in order (round-court)"]
    hdr(ws, _hdrs)
    ws.row_dimensions[1].height = 28

    for _ci, _w in enumerate([5, 14, 8, 10, 12, 10, 10, 80], 1):
        ws.column_dimensions[get_column_letter(_ci)].width = _w

    for _ri, _player in enumerate(players):
        _row = _ri + 2
        _pid = _player["id"]
        _g, _cap, _team = _player["gender"], _player["isCaptain"], _player["team"]
        _partner_id = _player["partnerId"]

        # Row fill
        if _cap and _team == "B":
            _rfill = R_PURP
        elif four_ids and _pid in four_ids:
            _rfill = R_ORNG
        elif _g == "F":
            _rfill = R_PINK
        else:
            _rfill = R_BLUE

        # Role
        if _cap:
            _role = "Captain"
        elif _partner_id:
            _role = "Couple"
        else:
            _role = "—"

        # Game list
        _pg = _player_games(_pid, pairs)
        _total = len(_pg)
        _mens  = sum(1 for _x in _pg if not _x[2])
        _mix   = sum(1 for _x in _pg if _x[2])

        _parts = []
        for (_r, _ct, _is_mix, _partner) in _pg:
            _pname = player_map[_partner]["name"]
            _parts.append(f"R{_r}-{_ct}{'★' if _is_mix else ''} w/{_pname}")
        _games_str = ",  ".join(_parts)

        _vals  = [_ri + 1, _player["name"], _g, _role, _total, _mens, _mix, _games_str]
        _bolds = [False, True, False, False, True, False, False, False]
        _aligns= ["center","left","center","center","center","center","center","left"]

        for _ci, (_v, _b, _al) in enumerate(zip(_vals, _bolds, _aligns), 1):
            _c = ws.cell(row=_row, column=_ci, value=_v)
            _c.fill  = _rfill
            _c.font  = Font(bold=_b)
            _c.alignment = Alignment(horizontal=_al, vertical="center",
                                      wrap_text=(_ci == 8))
            _c.border = border

        ws.row_dimensions[_row].height = 30 if len(_games_str) > 60 else 18

    ws.freeze_panes = "A2"
    return ws


# ════════════════════════════════════════════════════════════════════════════════
# 5. HSB Pairings Matrix
# ════════════════════════════════════════════════════════════════════════════════
write_matrix(
    title    = "Hill Street Blues — Player Pairings Matrix   (X = Men's Doubles  |  M★ = Mixed Doubles)",
    subtitle = "Captain: Suzan  ·  No fixed couples  ·  3 women × 1–3 mixed games  ·  9 men × 3–4 games each",
    players  = HSB_ORDER,
    pairs    = hsb_pairs,
    sheet_name = "HSB Pairings Matrix",
)

# ════════════════════════════════════════════════════════════════════════════════
# 6. HSB Player Summary
# ════════════════════════════════════════════════════════════════════════════════
write_summary("HSB Player Summary", HSB_ORDER, hsb_pairs)

# ════════════════════════════════════════════════════════════════════════════════
# 7. Blue Crew Pairings Matrix
# ════════════════════════════════════════════════════════════════════════════════
write_matrix(
    title    = "Blue Crew (JPL Team) — Pairings Matrix   (X = Men's Doubles  |  M★ = Mixed Doubles  |  C★ = Captain's game)",
    subtitle = "Captain: Cora (1 game w/ Jay)  ·  Couples: Trevor↔Alexis, Marv↔Carmela, Pierre↔Ivy  ·  No back-to-back games",
    players  = BC_ORDER,
    pairs    = bc_pairs,
    sheet_name = "Blue Crew Matrix",
    cap_pair = CAPTAIN_PAIR,
    four_ids = FOUR_GAME_IDS,
)

# ════════════════════════════════════════════════════════════════════════════════
# 8. Blue Crew Player Summary
# ════════════════════════════════════════════════════════════════════════════════
write_summary("Blue Crew Player Summary", BC_ORDER, bc_pairs, four_ids=FOUR_GAME_IDS)


wb.save("backend/ATH_Open_Sheets_Template.xlsx")
print("Saved: backend/ATH_Open_Sheets_Template.xlsx")
